"""Build data.json for the Foreign Property Screener.

The page is a market screener for a foreign investor: one row per country, every
factor a column, filterable. So this emits the whole workbook rather than the
handful of fields the old after-tax calculator read.

Two things are deliberately NOT carried through:

* The workbook's per-country ten-year growth column. Twenty-three of the
  thirty-four were "Est." with nothing behind them. Growth stays a reader input.
* The workbook's own 1-5 "foreign-buyer friendliness" score. It was a weighted
  composite of unsourced inputs with no working shown, which is exactly what a
  reader cannot check. The `ease` block below replaces it: the same idea built
  from six ordinals, each derived from a stated fact and each carried separately
  so the page can show WHY a market scores what it does rather than asserting it.

Destination tax rates come from rates_pwc.json where PwC has been read; the
workbook fills the rest and those rows are flagged unverified.

Run: python build_data.py
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

import classify

HERE = Path(__file__).parent
SRC = HERE / "source" / "International Property.xlsx"
OUT = HERE / "data.json"
PWC_FILE = HERE / "rates_pwc.json"

# Comparison sheet, header on row 2, data from row 3.
COL = {
    "country": 0, "price_local": 1, "currency": 2, "price_aud": 3, "fx_regime": 4,
    "price_to_income": 10, "gross_yield_centre": 11, "gross_yield": 12, "net_yield": 13,
    "population": 14, "pop_growth": 15, "urbanisation": 16, "gdp_per_capita": 17,
    "gdp_growth": 18, "inflation": 19, "sp_rating": 20, "property_rights": 21,
    "econ_freedom": 22, "ownership": 23, "visa": 24, "repatriation": 25,
    "cpi_score": 26, "purchase_costs": 27, "holding_costs": 28,
    "cgt_text": 29, "rental_tax_text": 30, "au_dta": 31, "wht_rent_text": 32,
    "estate_text": 33, "fx_vol": 34, "liquidity": 35, "obstacles": 36,
}


def clean(v):
    """Workbook cells carry non-breaking spaces and stray newlines."""
    if v is None:
        return ""
    s = unicodedata.normalize("NFKC", str(v)).replace("\xa0", " ")
    return re.sub(r"\s+", " ", s).strip()


def num(v):
    """First number in a cell, or the midpoint of a range like '3-6 months'."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = clean(v)
    rng = re.search(r"(\d+(?:\.\d+)?)\s*(?:-|to)\s*(\d+(?:\.\d+)?)", s)
    if rng:
        return round((float(rng.group(1)) + float(rng.group(2))) / 2, 2)
    one = re.search(r"(\d+(?:\.\d+)?)", s)
    return float(one.group(1)) if one else None


def yes_no(v):
    s = clean(v).lower()
    if s.startswith("yes"):
        return True
    if s.startswith("no"):
        return False
    return None


# Ease of investing is six ordinals, each 0 (hardest) to 3 (easiest), read off a
# stated fact and carried separately as well as summed, so the page can show WHY
# a market scores what it does. The prose-to-ordinal rules live in classify.py,
# which exists because the first version of them read three columns backwards.

def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    comp = wb["Comparison"]
    pwc = json.loads(PWC_FILE.read_text(encoding="utf-8"))
    profiles = {clean(r[0]): clean(r[1]) for r in wb["Country Profiles"].iter_rows(min_row=2, values_only=True) if r[0]}

    rows = []
    for r in comp.iter_rows(min_row=3, values_only=True):
        name = clean(r[COL["country"]])
        if not name:
            continue
        g = lambda k: r[COL[k]]

        purchase = num(g("purchase_costs"))
        months = num(g("liquidity"))
        rights = num(g("property_rights"))

        parts = {
            "ownership": classify.ownership(clean(g("ownership"))),
            "visa": classify.visa(clean(g("visa"))),
            "repatriation": classify.repatriation(clean(g("repatriation"))),
            "liquidity": classify.liquidity(months),
            "costs": classify.costs(purchase),
            "rights": classify.rights(rights),
        }

        got = [p[0] for p in parts.values() if p[0] is not None]
        ease = round(sum(got) / (3 * len(got)) * 100) if got else None

        p = pwc.get(name) or {}
        row = {
            "country": name,
            "currency": clean(g("currency")),
            "price_aud": g("price_aud"),
            "price_local": clean(g("price_local")),
            "fx_regime": clean(g("fx_regime")),
            "fx_vol": num(g("fx_vol")),
            "gross_yield": num(g("gross_yield")),
            "net_yield": num(g("net_yield")),
            "price_to_income": num(g("price_to_income")),
            "population": num(g("population")),
            "pop_growth": num(g("pop_growth")),
            "urbanisation": num(g("urbanisation")),
            "gdp_per_capita": num(g("gdp_per_capita")),
            "gdp_growth": num(g("gdp_growth")),
            "inflation": num(g("inflation")),
            "sp_rating": clean(g("sp_rating")),
            "property_rights": rights,
            "econ_freedom": num(g("econ_freedom")),
            "cpi_score": num(g("cpi_score")),
            "purchase_costs": purchase,
            "holding_costs": num(g("holding_costs")),
            "months_to_sell": months,
            "months_text": clean(g("liquidity")).replace(" months", "").split(" (")[0],
            "rental_tax_text": clean(g("rental_tax_text")),
            "cgt_text": clean(g("cgt_text")),
            "estate_text": clean(g("estate_text")),
            "wht_rent_text": clean(g("wht_rent_text")),
            "au_dta": yes_no(g("au_dta")),
            "ownership": clean(g("ownership")),
            "visa": clean(g("visa")),
            "repatriation": clean(g("repatriation")),
            "obstacles": clean(g("obstacles")),
            "profile": profiles.get(name, ""),
            "ease": ease,
            "ease_parts": {k: {"score": v[0], "label": v[1]} for k, v in parts.items()},
            "verified": bool(p),
            "rent_rate": (p.get("rent") or {}).get("rate"),
            "rent_basis": (p.get("rent") or {}).get("basis"),
            "rent_note": (p.get("rent") or {}).get("note", ""),
            "cgt_rate": (p.get("cgt") or {}).get("rate"),
            "cgt_basis": (p.get("cgt") or {}).get("basis"),
            "cgt_note": (p.get("cgt") or {}).get("note", ""),
        }
        rows.append(row)

    rows.sort(key=lambda d: d["country"])

    USED = {
        "Price-to-Income, Gross Rental Yields", "Price per sqm (city centre)",
        "Converted Price (AUD)", "Non-Resident Tax Rates (CGT, Rental, WHT)",
        "Australia Double Tax Agreements", "Purchase/Holding Costs",
        "Property Rights Score", "Economic Freedom Score", "CPI Score (Corruption)",
        "S&P Sovereign Rating", "GDP Growth Forecast",
        "Population, Urbanisation, GDP/Capita PPP", "FX Regimes",
        "Foreign Ownership Restrictions", "Golden Visa / Residency Pathways",
        "Estate/Inheritance Tax",
    }
    sources = []
    for r in wb["Sources"].iter_rows(min_row=2, values_only=True):
        if r[0] and clean(r[0]) in USED:
            sources.append({
                "measure": clean(r[0]), "name": clean(r[1]),
                "url": clean(r[2]), "caveat": clean(r[4]) if len(r) > 4 else "",
            })

    OUT.write_text(json.dumps({"countries": rows, "sources": sources}, indent=1, ensure_ascii=False), encoding="utf-8")
    print(f"{len(rows)} markets, {len(sources)} sources -> {OUT.name} ({OUT.stat().st_size // 1024}KB)")

    missing = [c["country"] for c in rows if c["ease"] is None]
    if missing:
        print("no ease score for:", ", ".join(missing), file=sys.stderr)


if __name__ == "__main__":
    main()
